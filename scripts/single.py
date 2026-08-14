#!/usr/bin/env python3
"""single.py

Research-facing post-analysis for a VerseVAD single-poem Complete Audit ZIP.

The program does not replace VerseVAD's measurement modules. It reads the
Complete Audit, uses the exported profile-aware metrics as authoritative, and
organizes/derives a close-reading package around one poem.

Core questions
--------------
1. What does this poem look like under one clearly stated primary profile?
2. How sensitive are those measurements to reasonable scope/weight/resource
   choices already represented in the Complete Audit?
3. Where do selected lexical measurements move through the poem?
4. Which lexical items contribute most strongly to those measurements?
5. Does any single line or stanza exert unusual influence on a whole-poem mean?
6. What supporting evidence about sound, form, readability, and coverage should
   remain visible beside the lexical analysis?

Expected input
--------------
A VerseVAD Single Poem Complete Audit ZIP. Export schema 3.0 is preferred; the script is
schema-aware and skips optional analyses when their supporting files are absent.

Primary outputs
---------------
exports/single_poem/<poem>_<timestamp>/
    single_poem_analysis.xlsx
    metric_summary.csv
    sensitivity_summary.csv
    sensitivity_variants.csv
    line_metrics.csv
    stanza_metrics.csv
    rolling_summary.csv
    rolling_windows.csv
    lexical_contributors.csv
    line_influence.csv
    stanza_influence.csv
    coverage_summary.csv
    structure_sound_form.csv
    readability_summary.csv
    additional_module_summary.csv
    analysis_spec.json
    analysis_metadata.json

Dependencies
------------
Python 3.10+, pandas, numpy, openpyxl.
"""

from __future__ import annotations

import argparse
import itertools
import json
import math
import re
import sys
import zipfile
from dataclasses import dataclass, asdict
from datetime import datetime
from functools import lru_cache
from pathlib import Path
from typing import Iterable, Optional, Sequence

try:
    import numpy as np
    import pandas as pd
except ImportError as exc:  # pragma: no cover
    raise SystemExit(
        "single.py requires numpy and pandas.\n"
        "Install them with:\n  python -m pip install numpy pandas openpyxl"
    ) from exc

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.dimensions import ColumnDimension
except ImportError as exc:  # pragma: no cover
    raise SystemExit(
        "single.py requires openpyxl for Excel export.\n"
        "Install it with:\n  python -m pip install openpyxl"
    ) from exc

from versevad_tools.cli import parse_index_selection
from versevad_tools.core import configure_console_encoding
from versevad_tools.audit import AuditSourceError, require_audit, resolve_member


__version__ = "0.1.0"

PROFILE_ORDER = [
    "ALL_LEXICAL__TOKEN",
    "ALL_LEXICAL__TYPE",
    "STOPWORD_EXCLUDED__TOKEN",
    "STOPWORD_EXCLUDED__TYPE",
    "CONTENT_WORDS__TOKEN",
    "CONTENT_WORDS__TYPE",
]

PROFILE_LABELS = {
    "ALL_LEXICAL__TOKEN": "All lexical tokens · Token-weighted",
    "ALL_LEXICAL__TYPE": "All lexical tokens · Type-weighted",
    "STOPWORD_EXCLUDED__TOKEN": "Stopword-excluded · Token-weighted",
    "STOPWORD_EXCLUDED__TYPE": "Stopword-excluded · Type-weighted",
    "CONTENT_WORDS__TOKEN": "Content words only · Token-weighted",
    "CONTENT_WORDS__TYPE": "Content words only · Type-weighted",
}

PROFILE_TO_SCOPE_WEIGHT = {
    "ALL_LEXICAL__TOKEN": ("ALL_LEXICAL", "TOKEN"),
    "ALL_LEXICAL__TYPE": ("ALL_LEXICAL", "TYPE"),
    "STOPWORD_EXCLUDED__TOKEN": ("STOPWORD_EXCLUDED", "TOKEN"),
    "STOPWORD_EXCLUDED__TYPE": ("STOPWORD_EXCLUDED", "TYPE"),
    "CONTENT_WORDS__TOKEN": ("CONTENT_WORDS", "TOKEN"),
    "CONTENT_WORDS__TYPE": ("CONTENT_WORDS", "TYPE"),
}

VAD_RESOURCES = {
    "warriner_vad_2013": "Warriner VAD",
    "nrc_vad_v1": "NRC VAD v1",
    "nrc_vad_v2_1": "NRC VAD v2.1",
}

# VerseVAD's profile layer preserves these words when the stopword-excluded
# view is requested. Keeping this explicit lets derived line/stanza/rolling
# calculations reproduce the profile table exactly for schema-v2 audits.
PROTECTED_STOPWORDS = {
    "against", "could", "least", "less", "may", "might", "more", "most",
    "must", "neither", "never", "no", "nor", "not", "should", "too",
    "very", "without",
}

# The schema-v2 profile comparison uses these four POS groups for the
# content-words view. Multiword affect matches are retained if any constituent
# token meets the content criterion, preserving the matched phrase as a unit.
CONTENT_POS = {"ADJ", "ADV", "NOUN", "VERB"}

PERCEPTUAL_SENSORIMOTOR = [
    "visual", "auditory", "haptic", "gustatory", "olfactory", "interoceptive"
]
ACTION_SENSORIMOTOR = ["foot_leg", "hand_arm", "head", "mouth", "torso"]
SENSORIMOTOR_ALL = PERCEPTUAL_SENSORIMOTOR + ACTION_SENSORIMOTOR

CORE_ORDER = [
    ("vad", "valence_mean"),
    ("vad", "arousal_mean"),
    ("vad", "dominance_mean"),
    ("concreteness", "concreteness_mean"),
    ("frequency", "frequency_mean"),
    ("aoa", "aoa_mean"),
] + [("sensorimotor", x) for x in SENSORIMOTOR_ALL]

DERIVABLE_CONTINUOUS_MODULES = {
    "vad", "concreteness", "frequency", "aoa", "sensorimotor",
    "emotion_intensity", "word_length"
}

OPTIONAL_SUMMARY_FILES = {
    "VADER": "01_AFFECT/vader/summary.csv",
    "Lexical Style": "03_LEXICAL_ACCESSIBILITY_AND_STYLE/lexical_style/summary.csv",
    "Pronunciation": "04_SOUND_AND_FORM/pronunciation/summary.csv",
    "Meter": "04_SOUND_AND_FORM/meter/summary.csv",
    "Rhyme & Sound": "04_SOUND_AND_FORM/rhyme_and_sound/summary.csv",
}

REQUIRED_FILES = [
    "00_START_HERE/profile_comparison.csv",
    "00_START_HERE/metric_dictionary.csv",
    "00_START_HERE/coverage_summary.csv",
    "07_PROCESSING_AUDIT/source.csv",
    "07_PROCESSING_AUDIT/tokens.csv",
]


@dataclass(frozen=True)
class MetricChoice:
    module_id: str
    metric_id: str
    source_id: str
    source: str
    metric: str
    unit: str
    measurement_kind: str = ""

    @property
    def key(self) -> str:
        return f"{self.module_id}|{self.metric_id}|{self.source_id}"

    @property
    def short_key(self) -> str:
        return f"{self.module_id}|{self.metric_id}"


@dataclass
class RunSpec:
    source_zip: str
    primary_profile: str
    primary_vad_resource: str
    selected_metrics: list[dict]
    sensitivity_profiles: list[str]
    sensitivity_vad_resources: list[str]
    rolling_windows: bool
    rolling_window_size: int
    rolling_step: int
    influence_analysis: bool
    top_contributors_per_metric: int


class AuditError(RuntimeError):
    pass


class SinglePoemAudit:
    """Reader/adapter for a VerseVAD single-poem Complete Audit ZIP."""

    def __init__(self, path: Path):
        self.path = Path(path).expanduser().resolve()
        if not self.path.exists():
            raise AuditError(f"File not found: {self.path}")
        if self.path.suffix.lower() != ".zip":
            raise AuditError("Expected a VerseVAD Complete Audit .zip file.")
        try:
            with zipfile.ZipFile(self.path) as zf:
                self.names = set(zf.namelist())
        except zipfile.BadZipFile as exc:
            raise AuditError(f"Not a readable ZIP archive: {self.path}") from exc
        try:
            self.descriptor = require_audit(
                self.path,
                expected_analysis_mode="single_poem",
                require_complete=True,
            )
        except AuditSourceError as exc:
            raise AuditError(str(exc)) from exc
        missing = [x for x in REQUIRED_FILES if resolve_member(self.names, x) is None]
        if missing:
            raise AuditError(
                "This does not look like a compatible VerseVAD single-poem Complete Audit.\n"
                "Missing required files:\n  " + "\n  ".join(missing)
            )

        self.profile = self.read_csv("00_START_HERE/profile_comparison.csv")
        self.metric_dictionary = self.read_csv("00_START_HERE/metric_dictionary.csv")
        self.coverage = self.read_csv("00_START_HERE/coverage_summary.csv")
        self.source = self.read_csv("07_PROCESSING_AUDIT/source.csv")
        self.tokens = self.read_csv("07_PROCESSING_AUDIT/tokens.csv")
        self._prepare_tokens()
        self.schema_version = self._detect_schema_version()

    @lru_cache(maxsize=128)
    def read_csv(self, name: str) -> pd.DataFrame:
        resolved = resolve_member(self.names, name)
        if resolved is None:
            raise AuditError(f"Archive does not contain {name}")
        with zipfile.ZipFile(self.path) as zf:
            with zf.open(resolved) as fh:
                return pd.read_csv(fh)

    def has(self, name: str) -> bool:
        return resolve_member(self.names, name) is not None

    def _detect_schema_version(self) -> Optional[int]:
        if not self.descriptor.legacy:
            try:
                return int(float(self.descriptor.schema_version))
            except ValueError:
                return None
        manifest = "08_REPRODUCIBILITY/master_manifest.csv"
        if manifest not in self.names:
            return None
        df = self.read_csv(manifest)
        if "audit_schema_version" not in df.columns or df.empty:
            return None
        vals = pd.to_numeric(df["audit_schema_version"], errors="coerce").dropna().unique()
        return int(vals[0]) if len(vals) == 1 else None

    def _prepare_tokens(self):
        t = self.tokens.copy()
        for c in ["is_lexical", "is_stopword", "is_punctuation", "is_proper_noun"]:
            if c in t.columns:
                t[c] = t[c].map(_as_bool)
        for c in ["token_position", "line_number", "stanza_number", "character_start", "character_end"]:
            if c in t.columns:
                t[c] = pd.to_numeric(t[c], errors="coerce")
        self.tokens = t
        self.token_by_id = t.set_index("token_id", drop=False)

    @property
    def title_raw(self) -> str:
        if self.source.empty:
            return self.path.stem
        return str(self.source.iloc[0].get("title", self.path.stem))

    @property
    def title(self) -> str:
        raw = self.title_raw.strip()
        m = re.match(r'^\s*["“]?(.+?)["”]?\s*\((.+?)\)\s*$', raw)
        return m.group(1).strip() if m else raw

    @property
    def author(self) -> str:
        raw = self.title_raw.strip()
        m = re.match(r'^\s*["“]?(.+?)["”]?\s*\((.+?)\)\s*$', raw)
        return m.group(2).strip() if m else ""

    @property
    def original_text(self) -> str:
        if self.source.empty:
            return ""
        v = self.source.iloc[0].get("original_text", "")
        return "" if pd.isna(v) else str(v)

    def evidence_availability(self) -> dict[str, bool]:
        return {
            "Profile-aware lexical metrics": self.has("00_START_HERE/profile_comparison.csv"),
            "Token-level processing audit": self.has("07_PROCESSING_AUDIT/tokens.csv"),
            "Line/stanza structure": self.has("07_PROCESSING_AUDIT/structure.csv"),
            "VAD / affect match evidence": self.has("01_AFFECT/lexical_match_audit.csv"),
            "Concreteness token evidence": self.has("02_EXPERIENCE_AND_IMAGERY/concreteness/token_audit.csv"),
            "Sensorimotor domain evidence": self.has("02_EXPERIENCE_AND_IMAGERY/sensorimotor/observations.csv"),
            "Frequency token evidence": self.has("03_LEXICAL_ACCESSIBILITY_AND_STYLE/frequency/token_audit.csv"),
            "AoA token evidence": self.has("03_LEXICAL_ACCESSIBILITY_AND_STYLE/age_of_acquisition/token_audit.csv"),
            "Sound/form evidence": any(self.has(p) for p in OPTIONAL_SUMMARY_FILES.values()),
            "Readability": self.has("03_LEXICAL_ACCESSIBILITY_AND_STYLE/readability/summary.csv"),
        }

    def vad_resources(self) -> list[tuple[str, str]]:
        p = self.profile
        ids = list(dict.fromkeys(p.loc[p["module_id"].eq("vad"), "source_id"].astype(str)))
        result = []
        for rid in ids:
            rows = p[p["source_id"].eq(rid)]
            source = str(rows.iloc[0]["source"]) if not rows.empty else VAD_RESOURCES.get(rid, rid)
            result.append((rid, source))
        return result

    def metric_catalog(self) -> list[MetricChoice]:
        cols = ["module_id", "source_id", "source", "metric_id", "metric", "unit"]
        c = self.profile[cols].drop_duplicates().copy()
        kind_map = {}
        if "measurement_kind" in self.metric_dictionary.columns:
            for _, row in self.metric_dictionary.iterrows():
                kind_map[(str(row.get("module_id", "")), str(row.get("metric_id", "")), str(row.get("source_id", "")))] = str(row.get("measurement_kind", ""))
        out = []
        for _, r in c.iterrows():
            key = (str(r.module_id), str(r.metric_id), str(r.source_id))
            out.append(MetricChoice(
                module_id=key[0], metric_id=key[1], source_id=key[2],
                source=str(r.source), metric=str(r.metric), unit=str(r.unit),
                measurement_kind=kind_map.get(key, "")
            ))
        return out

    def core_metrics(self, primary_vad_resource: str) -> list[MetricChoice]:
        catalog = self.metric_catalog()
        result = []
        for module, metric in CORE_ORDER:
            candidates = [x for x in catalog if x.module_id == module and x.metric_id == metric]
            if module == "vad":
                candidates = [x for x in candidates if x.source_id == primary_vad_resource]
            if candidates:
                result.append(candidates[0])
        return result

    def profile_row(self, choice: MetricChoice, profile_id: str) -> Optional[pd.Series]:
        p = self.profile
        rows = p[
            p["module_id"].astype(str).eq(choice.module_id)
            & p["metric_id"].astype(str).eq(choice.metric_id)
            & p["source_id"].astype(str).eq(choice.source_id)
            & p["profile_id"].astype(str).eq(profile_id)
        ]
        return None if rows.empty else rows.iloc[0]

    def _token_scope_mask(self, df: pd.DataFrame, scope: str, token_col: str = "token_id") -> pd.Series:
        meta = self.tokens[["token_id", "is_stopword", "normalized_form", "part_of_speech"]].copy()
        merged = df[[token_col]].merge(meta, left_on=token_col, right_on="token_id", how="left")
        merged.index = df.index
        if scope == "ALL_LEXICAL":
            return pd.Series(True, index=df.index)
        if scope == "STOPWORD_EXCLUDED":
            norm = merged["normalized_form"].fillna("").astype(str).str.lower()
            return (~merged["is_stopword"].fillna(False).astype(bool)) | norm.isin(PROTECTED_STOPWORDS)
        if scope == "CONTENT_WORDS":
            return merged["part_of_speech"].isin(CONTENT_POS)
        raise AuditError(f"Unknown scope {scope}")

    def _multi_token_scope_mask(self, df: pd.DataFrame, scope: str, token_ids_col: str = "token_ids") -> pd.Series:
        if scope == "ALL_LEXICAL":
            return pd.Series(True, index=df.index)
        token_meta = self.token_by_id

        def qualifies(raw) -> bool:
            ids = _split_token_ids(raw)
            flags = []
            for tid in ids:
                if tid not in token_meta.index:
                    continue
                row = token_meta.loc[tid]
                if scope == "STOPWORD_EXCLUDED":
                    ok = (not bool(row.get("is_stopword", False))) or str(row.get("normalized_form", "")).lower() in PROTECTED_STOPWORDS
                else:
                    ok = str(row.get("part_of_speech", "")) in CONTENT_POS
                flags.append(ok)
            # VerseVAD preserves a matched phrase when any token in that matched
            # phrase qualifies for the requested lexical scope.
            return any(flags)

        return df[token_ids_col].map(qualifies)

    def _apply_weighting(self, df: pd.DataFrame, weighting: str, group_cols: Optional[list[str]] = None) -> pd.DataFrame:
        if weighting == "TOKEN":
            return df
        if weighting != "TYPE":
            raise AuditError(f"Unknown weighting {weighting}")
        subset = ([*group_cols, "term_id"] if group_cols else ["term_id"])
        return df.drop_duplicates(subset=subset, keep="first")

    @lru_cache(maxsize=256)
    def observations(self, module_id: str, metric_id: str, source_id: str, profile_id: str) -> pd.DataFrame:
        """Return exact matched observations under one profile.

        Columns are normalized to:
          obs_id, token_position, line_number, stanza_number, term_id,
          term, surface, value, token_ids
        """
        scope, weighting = PROFILE_TO_SCOPE_WEIGHT[profile_id]

        if module_id == "vad":
            return self._obs_vad(metric_id, source_id, scope, weighting)
        if module_id == "concreteness":
            return self._obs_single_token(
                "02_EXPERIENCE_AND_IMAGERY/concreteness/token_audit.csv",
                "rating", scope, weighting
            )
        if module_id == "frequency":
            return self._obs_single_token(
                "03_LEXICAL_ACCESSIBILITY_AND_STYLE/frequency/token_audit.csv",
                "zipf_value", scope, weighting
            )
        if module_id == "aoa":
            return self._obs_single_token(
                "03_LEXICAL_ACCESSIBILITY_AND_STYLE/age_of_acquisition/token_audit.csv",
                "mean_age", scope, weighting
            )
        if module_id == "sensorimotor":
            return self._obs_sensorimotor(metric_id, scope, weighting)
        if module_id == "emotion_intensity":
            return self._obs_emotion_intensity(metric_id, scope, weighting)
        if module_id == "word_length":
            return self._obs_word_length(scope, weighting)
        raise AuditError(f"No exact observation adapter for module {module_id}")

    def _obs_single_token(self, path: str, value_col: str, scope: str, weighting: str) -> pd.DataFrame:
        if path not in self.names:
            return _empty_obs()
        d = self.read_csv(path).copy()
        d = d[d.get("included", False).map(_as_bool) if "included" in d.columns else pd.Series(True, index=d.index)]
        mask = self._token_scope_mask(d, scope, "token_id")
        d = d[mask].copy()
        d["value"] = pd.to_numeric(d[value_col], errors="coerce")
        d = d[d["value"].notna()]
        d["term_id"] = d.get("source_row", d.get("matched_lookup_form", d.get("normalized_form"))).astype(str)
        d["term"] = d.get("matched_lookup_form", d.get("normalized_form", d.get("surface_form"))).fillna(d.get("surface_form", "")).astype(str)
        d["surface"] = d.get("surface_form", d["term"]).astype(str)
        d["token_ids"] = d["token_id"].astype(str)
        d["obs_id"] = d["token_id"].astype(str)
        keep = ["obs_id", "token_position", "line_number", "stanza_number", "term_id", "term", "surface", "value", "token_ids"]
        d = d[keep].copy()
        return self._apply_weighting(d, weighting).reset_index(drop=True)

    def _obs_vad(self, metric_id: str, source_id: str, scope: str, weighting: str) -> pd.DataFrame:
        path = "01_AFFECT/lexical_match_audit.csv"
        if path not in self.names:
            return _empty_obs()
        value_map = {
            "valence_mean": "normalized_valence",
            "arousal_mean": "normalized_arousal",
            "dominance_mean": "normalized_dominance",
        }
        if metric_id not in value_map:
            return _empty_obs()
        d = self.read_csv(path).copy()
        d = d[d["lexicon_id"].astype(str).eq(source_id) & d["included"].map(_as_bool)].copy()
        mask = self._multi_token_scope_mask(d, scope, "token_ids")
        d = d[mask].copy()
        d["value"] = pd.to_numeric(d[value_map[metric_id]], errors="coerce")
        d = d[d["value"].notna()]
        d["term_id"] = d.get("source_rows", d.get("matched_lookup_form", d.get("matched_term"))).astype(str)
        d["term"] = d.get("matched_lookup_form", d.get("matched_term", d.get("normalized_span"))).fillna(d.get("normalized_span", "")).astype(str)
        d["surface"] = d.get("surface_span", d["term"]).fillna(d["term"]).astype(str)
        d["token_position"] = pd.to_numeric(d.get("start_token_position"), errors="coerce")
        d["obs_id"] = d.get("match_id", d.index.astype(str)).astype(str)
        keep = ["obs_id", "token_position", "line_number", "stanza_number", "term_id", "term", "surface", "value", "token_ids"]
        d = d[keep].copy()
        return self._apply_weighting(d, weighting).reset_index(drop=True)

    def _obs_sensorimotor(self, metric_id: str, scope: str, weighting: str) -> pd.DataFrame:
        path = "02_EXPERIENCE_AND_IMAGERY/sensorimotor/observations.csv"
        if path not in self.names:
            return _empty_obs()
        value_col = f"mean_{metric_id}"
        d = self.read_csv(path).copy()
        if value_col not in d.columns:
            return _empty_obs()
        # Schema-v2 Lancaster observations are single-token observations. Use
        # the processing audit's scope logic so the derived profile reproduces
        # profile_comparison.csv exactly.
        d["token_id"] = d["token_ids"].astype(str).str.split("|").str[0].str.strip()
        mask = self._token_scope_mask(d, scope, "token_id")
        d = d[mask].copy()
        d["value"] = pd.to_numeric(d[value_col], errors="coerce")
        d = d[d["value"].notna()]
        d["term_id"] = d.get("source_row", d.get("matched_lookup_form", d.get("normalized_surface"))).astype(str)
        d["term"] = d.get("matched_lookup_form", d.get("normalized_surface", d.get("surface_form"))).fillna(d.get("surface_form", "")).astype(str)
        d["surface"] = d.get("surface_form", d["term"]).astype(str)
        d["obs_id"] = d.get("observation_id", d.index.astype(str)).astype(str)
        keep = ["obs_id", "token_position", "line_number", "stanza_number", "term_id", "term", "surface", "value", "token_ids"]
        d = d[keep].copy()
        return self._apply_weighting(d, weighting).reset_index(drop=True)

    def _obs_emotion_intensity(self, metric_id: str, scope: str, weighting: str) -> pd.DataFrame:
        path = "01_AFFECT/lexical_match_audit.csv"
        if path not in self.names:
            return _empty_obs()
        emotion = metric_id.removesuffix("_intensity")
        d = self.read_csv(path).copy()
        d = d[d["lexicon_id"].astype(str).eq("nrc_emotion_intensity_v1") & d["included"].map(_as_bool)].copy()
        mask = self._multi_token_scope_mask(d, scope, "token_ids")
        d = d[mask].copy()

        def intensity_value(raw):
            if pd.isna(raw):
                return np.nan
            try:
                obj = json.loads(str(raw))
                return obj.get(emotion, np.nan)
            except Exception:
                return np.nan

        d["value"] = d["intensities"].map(intensity_value)
        d["value"] = pd.to_numeric(d["value"], errors="coerce")
        d = d[d["value"].notna()]
        d["term_id"] = d.get("source_rows", d.get("matched_lookup_form", d.get("matched_term"))).astype(str)
        d["term"] = d.get("matched_lookup_form", d.get("matched_term", d.get("normalized_span"))).fillna(d.get("normalized_span", "")).astype(str)
        d["surface"] = d.get("surface_span", d["term"]).fillna(d["term"]).astype(str)
        d["token_position"] = pd.to_numeric(d.get("start_token_position"), errors="coerce")
        d["obs_id"] = d.get("match_id", d.index.astype(str)).astype(str)
        keep = ["obs_id", "token_position", "line_number", "stanza_number", "term_id", "term", "surface", "value", "token_ids"]
        d = d[keep].copy()
        return self._apply_weighting(d, weighting).reset_index(drop=True)

    def _obs_word_length(self, scope: str, weighting: str) -> pd.DataFrame:
        path = "03_LEXICAL_ACCESSIBILITY_AND_STYLE/lexical_style/token_audit.csv"
        if path not in self.names:
            return _empty_obs()
        d = self.read_csv(path).copy()
        d = d[d["included"].map(_as_bool)].copy()
        mask = self._token_scope_mask(d, scope, "token_id")
        d = d[mask].copy()
        d["value"] = pd.to_numeric(d["alphabetic_character_count"], errors="coerce")
        d = d[d["value"].notna()]
        d["term_id"] = d.get("normalized_surface_type", d.get("surface_form")).astype(str)
        d["term"] = d["term_id"].astype(str)
        d["surface"] = d.get("surface_form", d["term"]).astype(str)
        d["token_ids"] = d["token_id"].astype(str)
        d["obs_id"] = d["token_id"].astype(str)
        keep = ["obs_id", "token_position", "line_number", "stanza_number", "term_id", "term", "surface", "value", "token_ids"]
        d = d[keep].copy()
        return self._apply_weighting(d, weighting).reset_index(drop=True)

    def association_contributors(self, metric_id: str, profile_id: str) -> pd.DataFrame:
        """Binary NRC Emotion Association contributors for one category."""
        path = "01_AFFECT/lexical_match_audit.csv"
        if path not in self.names:
            return pd.DataFrame()
        category = metric_id.removesuffix("_association")
        scope, weighting = PROFILE_TO_SCOPE_WEIGHT[profile_id]
        d = self.read_csv(path).copy()
        d = d[d["lexicon_id"].astype(str).eq("nrc_emotion_v0_92") & d["included"].map(_as_bool)].copy()
        mask = self._multi_token_scope_mask(d, scope, "token_ids")
        d = d[mask].copy()
        d["has_category"] = d["associations"].fillna("").astype(str).str.split("|").map(
            lambda xs: category in {x.strip().lower() for x in xs if x.strip()}
        )
        d = d[d["has_category"]].copy()
        if d.empty:
            return pd.DataFrame()
        d["term_id"] = d.get("source_rows", d.get("matched_lookup_form", d.get("matched_term"))).astype(str)
        if weighting == "TYPE":
            d = d.drop_duplicates("term_id")
        d["term"] = d.get("matched_lookup_form", d.get("matched_term", d.get("normalized_span"))).fillna(d.get("normalized_span", "")).astype(str)
        d["surface"] = d.get("surface_span", d["term"]).fillna(d["term"]).astype(str)
        rows = []
        for term_id, g in d.groupby("term_id", sort=False):
            rows.append({
                "term_id": term_id,
                "term": _first_nonempty(g["term"]),
                "occurrences": len(g),
                "surface_forms": " | ".join(sorted(set(g["surface"].dropna().astype(str)))),
                "lines": _join_ints(g["line_number"]),
                "stanzas": _join_ints(g["stanza_number"]),
                "contributor_kind": "binary association; no strength score",
            })
        return pd.DataFrame(rows)

    def eligible_token_positions(self, profile_id: str) -> pd.DataFrame:
        scope, _ = PROFILE_TO_SCOPE_WEIGHT[profile_id]
        d = self.tokens[self.tokens["is_lexical"].fillna(False).astype(bool)].copy()
        if scope == "STOPWORD_EXCLUDED":
            norm = d["normalized_form"].fillna("").astype(str).str.lower()
            d = d[(~d["is_stopword"].fillna(False).astype(bool)) | norm.isin(PROTECTED_STOPWORDS)]
        elif scope == "CONTENT_WORDS":
            d = d[d["part_of_speech"].isin(CONTENT_POS)]
        return d.sort_values("token_position").reset_index(drop=True)

    def structure_map(self, kind: str) -> pd.DataFrame:
        path = "07_PROCESSING_AUDIT/structure.csv"
        if path not in self.names:
            return pd.DataFrame()
        d = self.read_csv(path).copy()
        d = d[d["kind"].astype(str).eq(kind)].copy()
        return d

    def validate_observation_adapter(self, choice: MetricChoice, profile_id: str, tol: float = 1e-9) -> tuple[bool, float, float]:
        if choice.module_id not in DERIVABLE_CONTINUOUS_MODULES:
            return False, np.nan, np.nan
        row = self.profile_row(choice, profile_id)
        if row is None:
            return False, np.nan, np.nan
        try:
            obs = self.observations(choice.module_id, choice.metric_id, choice.source_id, profile_id)
        except AuditError:
            return False, np.nan, np.nan
        if obs.empty:
            return False, np.nan, float(row.get("value", np.nan))
        derived = float(obs["value"].mean())
        exported = float(row.get("value", np.nan))
        return bool(np.isfinite(derived) and np.isfinite(exported) and abs(derived - exported) <= tol), derived, exported


def _empty_obs() -> pd.DataFrame:
    return pd.DataFrame(columns=[
        "obs_id", "token_position", "line_number", "stanza_number",
        "term_id", "term", "surface", "value", "token_ids"
    ])


def _as_bool(v) -> bool:
    if isinstance(v, (bool, np.bool_)):
        return bool(v)
    if pd.isna(v):
        return False
    return str(v).strip().lower() in {"1", "true", "yes", "y", "t"}


def _split_token_ids(raw) -> list[str]:
    if pd.isna(raw):
        return []
    return [x.strip() for x in str(raw).split("|") if x.strip()]


def _first_nonempty(series: pd.Series) -> str:
    for v in series:
        if pd.notna(v) and str(v).strip():
            return str(v)
    return ""


def _join_ints(series: pd.Series) -> str:
    vals = sorted({int(x) for x in pd.to_numeric(series, errors="coerce").dropna() if int(x) > 0})
    return ", ".join(map(str, vals))


def safe_float(v) -> float:
    try:
        f = float(v)
        return f if math.isfinite(f) else np.nan
    except Exception:
        return np.nan


def mean_abs_pairwise(values: Sequence[float]) -> float:
    vals = [float(x) for x in values if pd.notna(x) and math.isfinite(float(x))]
    if len(vals) < 2:
        return np.nan
    diffs = [abs(a - b) for a, b in itertools.combinations(vals, 2)]
    return float(np.mean(diffs)) if diffs else np.nan


def slugify(text: str, max_len: int = 70) -> str:
    s = re.sub(r"[^A-Za-z0-9]+", "_", text).strip("_")
    return (s or "analysis")[:max_len]


def parse_multi_selection(raw: str, n: int, default: Optional[Sequence[int]] = None, allow_all: bool = True) -> list[int]:
    text = raw.strip()
    if not text and default is not None:
        return list(default)
    return parse_index_selection(text, n, allow_all=allow_all, one_based=True)


def prompt_multi(prompt: str, n: int, default: Optional[Sequence[int]] = None, allow_all: bool = True) -> list[int]:
    while True:
        try:
            return parse_multi_selection(input(prompt), n, default=default, allow_all=allow_all)
        except ValueError as exc:
            print(f"  {exc}")
            print("  Use A, 1,2, 1,3,5, 5-6, or combinations such as 1,3,5-6.")


def prompt_yes_no(prompt: str, default: bool = True) -> bool:
    suffix = " [Y/n]: " if default else " [y/N]: "
    while True:
        raw = input(prompt + suffix).strip().lower()
        if not raw:
            return default
        if raw in {"y", "yes"}:
            return True
        if raw in {"n", "no"}:
            return False
        print("  Enter y or n.")


def prompt_int(prompt: str, default: int, minimum: int = 1, maximum: Optional[int] = None) -> int:
    while True:
        raw = input(f"{prompt} [{default}]: ").strip()
        if not raw:
            return default
        try:
            v = int(raw)
        except ValueError:
            print("  Enter a whole number.")
            continue
        if v < minimum or (maximum is not None and v > maximum):
            print(f"  Enter a value from {minimum}" + (f" to {maximum}." if maximum else " or higher."))
            continue
        return v


def discover_source_zips(script_path: Path) -> tuple[Path, list[Path]]:
    script_dir = script_path.resolve().parent
    candidates = [script_dir.parent, Path.cwd().resolve()]
    seen = set()
    for root in candidates:
        if root in seen:
            continue
        seen.add(root)
        source_dir = root / "source"
        if source_dir.exists():
            zips = sorted(source_dir.glob("*.zip"), key=lambda p: p.name.lower())
            compatible = []
            for p in zips:
                try:
                    require_audit(
                        p,
                        expected_analysis_mode="single_poem",
                        require_complete=True,
                    )
                    compatible.append(p)
                except (zipfile.BadZipFile, AuditSourceError):
                    pass
            if compatible:
                return root, compatible
    return Path.cwd().resolve(), []


def choose_source(args, script_path: Path) -> tuple[Path, Path]:
    if args.source:
        p = Path(args.source).expanduser().resolve()
        root = p.parent.parent if p.parent.name == "source" else Path.cwd().resolve()
        return root, p
    root, files = discover_source_zips(script_path)
    if not files:
        raise SystemExit(
            "No compatible VerseVAD single-poem Complete Audit ZIPs were found in source/.\n"
            "Put an audit ZIP in the project's source folder or use --source PATH."
        )
    if len(files) == 1:
        print(f"Found one compatible VerseVAD single-poem audit:\n  {files[0].name}")
        if not args.quick and not prompt_yes_no("Use this audit?", True):
            raise SystemExit("Cancelled.")
        return root, files[0]
    print("Available VerseVAD single-poem audits:\n")
    for i, p in enumerate(files, 1):
        print(f"[{i}] {p.name}")
    if args.quick:
        return root, files[0]
    idx = prompt_multi("Select one: ", len(files), allow_all=False)[0]
    return root, files[idx - 1]


def choose_profile(args) -> str:
    if args.primary_profile:
        raw = args.primary_profile.strip().upper()
        if raw in PROFILE_ORDER:
            return raw
        if raw.isdigit() and 1 <= int(raw) <= len(PROFILE_ORDER):
            return PROFILE_ORDER[int(raw) - 1]
        raise SystemExit("Invalid --primary-profile. Use 1-6 or an exact profile ID.")
    print("\nPrimary lexical profile\n")
    for i, pid in enumerate(PROFILE_ORDER, 1):
        marker = " [default]" if pid == "STOPWORD_EXCLUDED__TOKEN" else ""
        print(f"[{i}] {PROFILE_LABELS[pid]}{marker}")
    if args.quick:
        return "STOPWORD_EXCLUDED__TOKEN"
    idx = prompt_multi("\nSelect one [3]: ", len(PROFILE_ORDER), default=[3], allow_all=False)[0]
    return PROFILE_ORDER[idx - 1]


def choose_vad_resource(audit: SinglePoemAudit, args) -> str:
    resources = audit.vad_resources()
    if not resources:
        return ""
    ids = [x[0] for x in resources]
    if args.primary_vad:
        if args.primary_vad in ids:
            return args.primary_vad
        raise SystemExit(f"Requested VAD resource is not present: {args.primary_vad}")
    default_id = "nrc_vad_v2_1" if "nrc_vad_v2_1" in ids else ids[0]
    default_index = ids.index(default_id) + 1
    print("\nPrimary VAD resource\n")
    for i, (rid, label) in enumerate(resources, 1):
        marker = " [default]" if i == default_index else ""
        print(f"[{i}] {label} ({rid}){marker}")
    if args.quick:
        return default_id
    idx = prompt_multi(f"\nSelect one [{default_index}]: ", len(resources), default=[default_index], allow_all=False)[0]
    return resources[idx - 1][0]


def metric_search_loop(audit: SinglePoemAudit, selected: list[MetricChoice], primary_vad: str) -> list[MetricChoice]:
    catalog = audit.metric_catalog()
    # VAD alternatives belong in the sensitivity-resource choice, not in the
    # main metric list, so search exposes only the chosen primary VAD resource.
    searchable = [m for m in catalog if m.module_id != "vad" or m.source_id == primary_vad]
    selected_keys = {x.key for x in selected}
    while True:
        query = input("\nSearch metric (blank to finish): ").strip()
        if not query:
            break
        q = query.lower()
        matches = [m for m in searchable if q in " ".join([m.module_id, m.metric_id, m.metric, m.source, m.unit]).lower() and m.key not in selected_keys]
        if not matches:
            print("  No unselected matching metrics.")
            continue
        print()
        for i, m in enumerate(matches[:30], 1):
            print(f"[{i}] {m.metric} · {m.source} [{m.unit}]")
        if len(matches) > 30:
            print(f"  Showing first 30 of {len(matches)} matches. Refine the search to narrow it.")
        picks = prompt_multi("Select one or more: ", min(len(matches), 30), allow_all=True)
        for i in picks:
            m = matches[i - 1]
            selected.append(m)
            selected_keys.add(m.key)
            print(f"  Added: {m.metric} · {m.source}")
    return selected


def build_metric_summary(audit: SinglePoemAudit, selected: list[MetricChoice], profile_id: str) -> pd.DataFrame:
    rows = []
    for m in selected:
        r = audit.profile_row(m, profile_id)
        if r is None:
            continue
        coverage = safe_float(r.get("token_coverage" if PROFILE_TO_SCOPE_WEIGHT[profile_id][1] == "TOKEN" else "type_coverage"))
        rows.append({
            "module": m.module_id,
            "metric_id": m.metric_id,
            "metric": m.metric,
            "resource_id": m.source_id,
            "resource": m.source,
            "profile_id": profile_id,
            "profile": PROFILE_LABELS[profile_id],
            "value": safe_float(r.get("value")),
            "median": safe_float(r.get("median")),
            "population_sd": safe_float(r.get("population_standard_deviation")),
            "q1": safe_float(r.get("first_quartile")),
            "q3": safe_float(r.get("third_quartile")),
            "minimum": safe_float(r.get("minimum")),
            "maximum": safe_float(r.get("maximum")),
            "observations": safe_float(r.get("observation_count")),
            "coverage": coverage,
            "unit": m.unit,
            "measurement_kind": m.measurement_kind,
        })
    return pd.DataFrame(rows)


def build_sensitivity(audit: SinglePoemAudit, selected: list[MetricChoice], primary_profile: str,
                      sensitivity_profiles: list[str], vad_resources: list[str]) -> tuple[pd.DataFrame, pd.DataFrame]:
    variant_rows = []
    summary_rows = []
    vad_lookup = {rid: label for rid, label in audit.vad_resources()}

    for m in selected:
        resources = vad_resources if m.module_id == "vad" else [m.source_id]
        metric_variants = []
        for rid in resources:
            source_label = vad_lookup.get(rid, m.source if rid == m.source_id else rid)
            temp = MetricChoice(m.module_id, m.metric_id, rid, source_label, m.metric, m.unit, m.measurement_kind)
            for pid in sensitivity_profiles:
                r = audit.profile_row(temp, pid)
                if r is None:
                    continue
                val = safe_float(r.get("value"))
                if not np.isfinite(val):
                    continue
                row = {
                    "module": m.module_id,
                    "metric_id": m.metric_id,
                    "metric": m.metric,
                    "resource_id": rid,
                    "resource": source_label,
                    "profile_id": pid,
                    "profile": PROFILE_LABELS[pid],
                    "value": val,
                    "observations": safe_float(r.get("observation_count")),
                    "token_coverage": safe_float(r.get("token_coverage")),
                    "type_coverage": safe_float(r.get("type_coverage")),
                    "unit": str(r.get("unit", m.unit)),
                    "is_primary": bool(rid == m.source_id and pid == primary_profile),
                }
                variant_rows.append(row)
                metric_variants.append(row)
        if not metric_variants:
            continue
        vals = [x["value"] for x in metric_variants]
        lo = min(metric_variants, key=lambda x: x["value"])
        hi = max(metric_variants, key=lambda x: x["value"])
        primary = next((x for x in metric_variants if x["is_primary"]), None)
        summary_rows.append({
            "module": m.module_id,
            "metric_id": m.metric_id,
            "metric": m.metric,
            "primary_resource": m.source,
            "primary_profile": PROFILE_LABELS[primary_profile],
            "primary_value": primary["value"] if primary else np.nan,
            "variants_compared": len(metric_variants),
            "lowest_value": lo["value"],
            "lowest_variant": f"{lo['resource']} · {lo['profile']}",
            "highest_value": hi["value"],
            "highest_variant": f"{hi['resource']} · {hi['profile']}",
            "maximum_absolute_change": hi["value"] - lo["value"],
            "average_absolute_pairwise_change": mean_abs_pairwise(vals),
            "unit": m.unit,
        })
    return pd.DataFrame(summary_rows), pd.DataFrame(variant_rows)


def build_unit_dynamics(audit: SinglePoemAudit, selected: list[MetricChoice], profile_id: str, unit: str) -> pd.DataFrame:
    unit_col = "line_number" if unit == "line" else "stanza_number"
    rows = []
    structure = audit.structure_map(unit)
    text_lookup = {}
    if not structure.empty:
        for _, r in structure.iterrows():
            text_lookup[int(r["ordinal"])] = str(r.get("content_text", r.get("raw_text", "")))

    for m in selected:
        if m.module_id not in DERIVABLE_CONTINUOUS_MODULES:
            continue
        ok, derived, exported = audit.validate_observation_adapter(m, profile_id)
        if not ok:
            continue
        obs = audit.observations(m.module_id, m.metric_id, m.source_id, profile_id)
        if obs.empty:
            continue
        _, weighting = PROFILE_TO_SCOPE_WEIGHT[profile_id]
        for ordinal, g in obs.groupby(unit_col):
            if pd.isna(ordinal) or int(ordinal) <= 0:
                continue
            local = g if weighting == "TOKEN" else g.drop_duplicates("term_id")
            if local.empty:
                continue
            rows.append({
                "unit": unit,
                "ordinal": int(ordinal),
                "source_text": text_lookup.get(int(ordinal), ""),
                "module": m.module_id,
                "metric_id": m.metric_id,
                "metric": m.metric,
                "resource": m.source,
                "profile": PROFILE_LABELS[profile_id],
                "value": float(local["value"].mean()),
                "observations": len(local),
                "unit_or_scale": m.unit,
            })
    return pd.DataFrame(rows)


def build_rolling_windows(audit: SinglePoemAudit, selected: list[MetricChoice], profile_id: str,
                          window_size: int, step: int) -> tuple[pd.DataFrame, pd.DataFrame]:
    eligible = audit.eligible_token_positions(profile_id)
    if len(eligible) < max(5, window_size):
        return pd.DataFrame(), pd.DataFrame()
    original_text = audit.original_text
    metric_obs = {}
    validated = []
    for m in selected:
        if m.module_id not in DERIVABLE_CONTINUOUS_MODULES:
            continue
        ok, _, _ = audit.validate_observation_adapter(m, profile_id)
        if ok:
            metric_obs[m.key] = audit.observations(m.module_id, m.metric_id, m.source_id, profile_id)
            validated.append(m)

    rows = []
    starts = list(range(0, max(1, len(eligible) - window_size + 1), step))
    if starts and starts[-1] != len(eligible) - window_size:
        starts.append(len(eligible) - window_size)
    for wi, start in enumerate(starts, 1):
        chunk = eligible.iloc[start:start + window_size]
        if chunk.empty:
            continue
        token_ids = set(chunk["token_id"].astype(str))
        pos_min = int(chunk["token_position"].min())
        pos_max = int(chunk["token_position"].max())
        char_start = int(chunk["character_start"].min()) if "character_start" in chunk and chunk["character_start"].notna().any() else None
        char_end = int(chunk["character_end"].max()) if "character_end" in chunk and chunk["character_end"].notna().any() else None
        snippet = ""
        if original_text and char_start is not None and char_end is not None:
            snippet = original_text[char_start:char_end].replace("\n", " ").strip()
        line_start = int(chunk["line_number"].dropna().min()) if chunk["line_number"].notna().any() else np.nan
        line_end = int(chunk["line_number"].dropna().max()) if chunk["line_number"].notna().any() else np.nan
        stanza_start = int(chunk["stanza_number"].dropna().min()) if chunk["stanza_number"].notna().any() else np.nan
        stanza_end = int(chunk["stanza_number"].dropna().max()) if chunk["stanza_number"].notna().any() else np.nan

        for m in validated:
            obs = metric_obs[m.key]
            # Include a phrase observation if any constituent token falls in the
            # current eligible-token window. Single-token observations are a
            # degenerate case of the same rule.
            mask = obs["token_ids"].map(lambda raw: bool(token_ids.intersection(_split_token_ids(raw))))
            local = obs[mask].copy()
            if local.empty:
                val = np.nan
                nobs = 0
            else:
                _, weighting = PROFILE_TO_SCOPE_WEIGHT[profile_id]
                if weighting == "TYPE":
                    local = local.drop_duplicates("term_id")
                val = float(local["value"].mean()) if not local.empty else np.nan
                nobs = len(local)
            rows.append({
                "window": wi,
                "eligible_token_start_index": start + 1,
                "eligible_token_end_index": start + len(chunk),
                "token_position_start": pos_min,
                "token_position_end": pos_max,
                "line_start": line_start,
                "line_end": line_end,
                "stanza_start": stanza_start,
                "stanza_end": stanza_end,
                "window_text": snippet,
                "module": m.module_id,
                "metric_id": m.metric_id,
                "metric": m.metric,
                "resource": m.source,
                "profile": PROFILE_LABELS[profile_id],
                "value": val,
                "observations": nobs,
                "unit_or_scale": m.unit,
            })

    long = pd.DataFrame(rows)
    if long.empty:
        return long, pd.DataFrame()
    summaries = []
    for (module, metric_id, resource), g in long.groupby(["module", "metric_id", "resource"], dropna=False):
        gg = g[g["value"].notna()].sort_values("window")
        if gg.empty:
            continue
        lo = gg.loc[gg["value"].idxmin()]
        hi = gg.loc[gg["value"].idxmax()]
        first = gg.iloc[0]
        last = gg.iloc[-1]
        diffs = gg["value"].diff()
        if diffs.notna().any():
            idx = diffs.abs().idxmax()
            jump = gg.loc[idx]
            previous = gg.loc[gg.index[gg.index.get_loc(idx)-1]] if gg.index.get_loc(idx) > 0 else None
            largest_change = float(diffs.loc[idx])
            change_from_window = int(previous["window"]) if previous is not None else np.nan
            change_to_window = int(jump["window"])
        else:
            largest_change = np.nan
            change_from_window = np.nan
            change_to_window = np.nan
        summaries.append({
            "module": module,
            "metric_id": metric_id,
            "metric": str(gg.iloc[0]["metric"]),
            "resource": resource,
            "profile": str(gg.iloc[0]["profile"]),
            "first_window_value": float(first["value"]),
            "last_window_value": float(last["value"]),
            "end_minus_beginning": float(last["value"] - first["value"]),
            "minimum_window_value": float(lo["value"]),
            "minimum_window": int(lo["window"]),
            "minimum_window_lines": f"{int(lo['line_start'])}-{int(lo['line_end'])}" if pd.notna(lo["line_start"]) else "",
            "maximum_window_value": float(hi["value"]),
            "maximum_window": int(hi["window"]),
            "maximum_window_lines": f"{int(hi['line_start'])}-{int(hi['line_end'])}" if pd.notna(hi["line_start"]) else "",
            "largest_adjacent_window_change": largest_change,
            "change_from_window": change_from_window,
            "change_to_window": change_to_window,
            "unit_or_scale": str(gg.iloc[0]["unit_or_scale"]),
        })
    return long, pd.DataFrame(summaries)


def build_contributors(audit: SinglePoemAudit, selected: list[MetricChoice], profile_id: str,
                       top_n: int) -> pd.DataFrame:
    rows = []
    for m in selected:
        if m.module_id == "emotion_association":
            d = audit.association_contributors(m.metric_id, profile_id)
            if d.empty:
                continue
            d = d.sort_values(["occurrences", "term"], ascending=[False, True]).head(top_n)
            for rank, (_, r) in enumerate(d.iterrows(), 1):
                rows.append({
                    "module": m.module_id, "metric_id": m.metric_id, "metric": m.metric,
                    "resource": m.source, "rank": rank, "term": r["term"],
                    "occurrences": r["occurrences"], "term_value": np.nan,
                    "difference_from_poem_mean": np.nan, "weighted_deviation_sum": np.nan,
                    "absolute_weighted_deviation": np.nan,
                    "surface_forms": r["surface_forms"], "lines": r["lines"], "stanzas": r["stanzas"],
                    "contributor_kind": r["contributor_kind"], "unit_or_scale": m.unit,
                })
            continue
        if m.module_id not in DERIVABLE_CONTINUOUS_MODULES:
            continue
        ok, _, exported = audit.validate_observation_adapter(m, profile_id)
        if not ok:
            continue
        obs = audit.observations(m.module_id, m.metric_id, m.source_id, profile_id)
        if obs.empty:
            continue
        poem_mean = float(exported)
        _, weighting = PROFILE_TO_SCOPE_WEIGHT[profile_id]
        for term_id, g in obs.groupby("term_id", sort=False):
            term_value = float(g["value"].mean())
            occurrences = len(g)
            effective_weight = occurrences if weighting == "TOKEN" else 1
            dev = effective_weight * (term_value - poem_mean)
            rows.append({
                "module": m.module_id,
                "metric_id": m.metric_id,
                "metric": m.metric,
                "resource": m.source,
                "rank": np.nan,
                "term": _first_nonempty(g["term"]),
                "occurrences": occurrences,
                "term_value": term_value,
                "difference_from_poem_mean": term_value - poem_mean,
                "weighted_deviation_sum": dev,
                "absolute_weighted_deviation": abs(dev),
                "surface_forms": " | ".join(sorted(set(g["surface"].dropna().astype(str)))),
                "lines": _join_ints(g["line_number"]),
                "stanzas": _join_ints(g["stanza_number"]),
                "contributor_kind": "continuous lexical contributor",
                "unit_or_scale": m.unit,
            })
    out = pd.DataFrame(rows)
    if out.empty:
        return out
    final = []
    for _, g in out.groupby(["module", "metric_id", "resource"], dropna=False):
        if g["absolute_weighted_deviation"].notna().any():
            gg = g.sort_values(["absolute_weighted_deviation", "occurrences"], ascending=[False, False]).head(top_n).copy()
        else:
            gg = g.sort_values(["occurrences", "term"], ascending=[False, True]).head(top_n).copy()
        gg["rank"] = range(1, len(gg) + 1)
        final.append(gg)
    return pd.concat(final, ignore_index=True) if final else pd.DataFrame()


def _recompute_after_removal(obs: pd.DataFrame, unit_col: str, unit_value: int, weighting: str) -> tuple[float, int]:
    remaining = obs[pd.to_numeric(obs[unit_col], errors="coerce").fillna(-1).astype(int) != int(unit_value)].copy()
    if weighting == "TYPE":
        remaining = remaining.drop_duplicates("term_id")
    if remaining.empty:
        return np.nan, 0
    return float(remaining["value"].mean()), len(remaining)


def build_influence(audit: SinglePoemAudit, selected: list[MetricChoice], profile_id: str, unit: str) -> pd.DataFrame:
    unit_col = "line_number" if unit == "line" else "stanza_number"
    structure = audit.structure_map(unit)
    text_lookup = {}
    if not structure.empty:
        text_lookup = {int(r["ordinal"]): str(r.get("content_text", r.get("raw_text", ""))) for _, r in structure.iterrows()}
    rows = []
    _, weighting = PROFILE_TO_SCOPE_WEIGHT[profile_id]
    for m in selected:
        if m.module_id not in DERIVABLE_CONTINUOUS_MODULES:
            continue
        ok, derived, exported = audit.validate_observation_adapter(m, profile_id)
        if not ok:
            continue
        obs = audit.observations(m.module_id, m.metric_id, m.source_id, profile_id)
        units = sorted({int(x) for x in pd.to_numeric(obs[unit_col], errors="coerce").dropna() if int(x) > 0})
        for ordinal in units:
            loo, n_remaining = _recompute_after_removal(obs, unit_col, ordinal, weighting)
            if not np.isfinite(loo):
                continue
            rows.append({
                "unit": unit,
                "ordinal": ordinal,
                "source_text": text_lookup.get(ordinal, ""),
                "module": m.module_id,
                "metric_id": m.metric_id,
                "metric": m.metric,
                "resource": m.source,
                "profile": PROFILE_LABELS[profile_id],
                "original_value": exported,
                "loo_value": loo,
                "delta": loo - exported,
                "absolute_delta": abs(loo - exported),
                "remaining_observations": n_remaining,
                "adapter_validation_error": derived - exported,
                "unit_or_scale": m.unit,
            })
    out = pd.DataFrame(rows)
    if out.empty:
        return out
    out["influence_rank"] = out.groupby(["module", "metric_id", "resource"])["absolute_delta"].rank(method="first", ascending=False)
    return out.sort_values(["module", "metric_id", "resource", "influence_rank", "ordinal"]).reset_index(drop=True)


def build_structure_sound_form(audit: SinglePoemAudit) -> pd.DataFrame:
    rows = []
    for label, path in OPTIONAL_SUMMARY_FILES.items():
        if label == "Lexical Style" or not audit.has(path):
            continue
        d = audit.read_csv(path)
        if not {"section", "metric", "value"}.issubset(d.columns):
            continue
        for _, r in d.iterrows():
            rows.append({
                "module": label,
                "section": r.get("section", ""),
                "metric": r.get("metric", ""),
                "value": r.get("value", ""),
                "unit_or_scale": r.get("unit_or_scale", ""),
                "denominator": r.get("denominator", ""),
                "note": r.get("note", ""),
            })
    path = "04_SOUND_AND_FORM/inherited_form/summary.csv"
    if audit.has(path):
        d = audit.read_csv(path)
        for _, r in d.iterrows():
            for metric in [
                "result_status", "best_candidate_name", "classification", "consistency_index",
                "evidence_coverage", "confidence_band", "nearest_alternative_name", "candidate_margin",
                "narrative"
            ]:
                if metric in d.columns:
                    rows.append({
                        "module": "Inherited Form", "section": "candidate", "metric": metric,
                        "value": r.get(metric, ""), "unit_or_scale": "", "denominator": "", "note": ""
                    })
    return pd.DataFrame(rows)


def build_readability(audit: SinglePoemAudit) -> pd.DataFrame:
    path = "03_LEXICAL_ACCESSIBILITY_AND_STYLE/readability/summary.csv"
    if not audit.has(path):
        return pd.DataFrame()
    return audit.read_csv(path).copy()


def build_additional_module_summary(audit: SinglePoemAudit, primary_profile: str, primary_vad: str) -> pd.DataFrame:
    rows = []
    for label, path in [
        ("VADER", "01_AFFECT/vader/summary.csv"),
        ("Lexical Style", "03_LEXICAL_ACCESSIBILITY_AND_STYLE/lexical_style/summary.csv"),
    ]:
        if audit.has(path):
            d = audit.read_csv(path)
            if {"section", "metric", "value"}.issubset(d.columns):
                for _, r in d.iterrows():
                    rows.append({"module": label, "section": r.get("section", ""), "metric": r.get("metric", ""), "value": r.get("value", ""), "unit_or_scale": r.get("unit_or_scale", ""), "denominator": r.get("denominator", ""), "note": r.get("note", "")})
    # PoetryID is included only as an interpretive profile summary. Affinities
    # and confidence labels are not probabilities.
    path = "05_COMPARATIVE_PROFILES/poetry_id/summary.csv"
    if audit.has(path):
        d = audit.read_csv(path)
        view_map = {
            "ALL_LEXICAL": "all_matched",
            "STOPWORD_EXCLUDED": "stopwords_excluded",
            "CONTENT_WORDS": "content_words",
        }
        scope, weighting = PROFILE_TO_SCOPE_WEIGHT[primary_profile]
        rows_pi = d[
            d["source_lexicon_id"].astype(str).eq(primary_vad)
            & d["analysis_view"].astype(str).eq(view_map[scope])
            & d["weighting_mode"].astype(str).str.upper().eq(weighting)
        ]
        if not rows_pi.empty:
            r = rows_pi.iloc[0]
            for metric in ["categorical_archetype_name", "short_descriptor", "nearest_centroid_archetype_name", "confidence_label", "coverage_assessment", "narrative_summary", "interpretive_caution"]:
                if metric in r.index:
                    rows.append({"module": "PoetryID", "section": "selected profile", "metric": metric, "value": r.get(metric, ""), "unit_or_scale": "interpretive label / narrative", "denominator": "", "note": "Affinity/nearest-centroid evidence is not a probability."})
    return pd.DataFrame(rows)


def influence_summary(influence: pd.DataFrame) -> pd.DataFrame:
    if influence.empty:
        return pd.DataFrame()
    rows = []
    for keys, g in influence.groupby(["unit", "module", "metric_id", "metric", "resource", "profile", "unit_or_scale"], dropna=False):
        best = g.loc[g["absolute_delta"].idxmax()]
        rows.append({
            "unit": keys[0], "module": keys[1], "metric_id": keys[2], "metric": keys[3],
            "resource": keys[4], "profile": keys[5], "unit_or_scale": keys[6],
            "original_value": float(best["original_value"]),
            "loo_minimum": float(g["loo_value"].min()),
            "loo_maximum": float(g["loo_value"].max()),
            "largest_absolute_change": float(best["absolute_delta"]),
            "most_influential_ordinal": int(best["ordinal"]),
            "most_influential_text": best.get("source_text", ""),
            "value_without_most_influential": float(best["loo_value"]),
            "signed_change": float(best["delta"]),
            "units_tested": len(g),
        })
    return pd.DataFrame(rows)


def save_csv(df: pd.DataFrame, path: Path):
    if df is None:
        df = pd.DataFrame()
    df.to_csv(path, index=False, encoding="utf-8-sig")


# ---------- Excel helpers ----------
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
SUBHEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
LIGHT_FILL = PatternFill("solid", fgColor="F3F6F9")
WHITE_FONT = Font(color="FFFFFF", bold=True)
BOLD_FONT = Font(bold=True)
THIN_GRAY = Side(style="thin", color="D9D9D9")
CELL_BORDER = Border(bottom=THIN_GRAY)


def _xlsx_safe(v):
    if isinstance(v, (np.integer,)):
        return int(v)
    if isinstance(v, (np.floating,)):
        return None if not np.isfinite(v) else float(v)
    if isinstance(v, float) and not math.isfinite(v):
        return None
    if pd.isna(v) if not isinstance(v, (list, dict, tuple, set)) else False:
        return None
    return v


def style_table(ws, freeze: str = "A2", autofilter: bool = True):
    if ws.max_row >= 1:
        for cell in ws[1]:
            cell.fill = HEADER_FILL
            cell.font = WHITE_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[1].height = 30
        if autofilter and ws.max_column > 0:
            ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = freeze
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = CELL_BORDER
    _autosize(ws)


def _autosize(ws, max_width: int = 45):
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        width = 10
        for cell in ws[letter]:
            val = "" if cell.value is None else str(cell.value)
            # Long poem text should not explode the sheet width.
            width = max(width, min(len(val) + 2, max_width))
        ws.column_dimensions[letter].width = min(width, max_width)


def write_df_sheet(wb: Workbook, name: str, df: pd.DataFrame, freeze: str = "A2"):
    ws = wb.create_sheet(_safe_sheet_name(name))
    if df is None or df.empty:
        ws.append(["No data available for this analysis."])
        ws["A1"].font = BOLD_FONT
        ws.column_dimensions["A"].width = 60
        return ws
    ws.append(list(df.columns))
    for row in df.itertuples(index=False, name=None):
        ws.append([_xlsx_safe(v) for v in row])
    style_table(ws, freeze=freeze)
    return ws


def _safe_sheet_name(name: str) -> str:
    name = re.sub(r"[\\/*?:\[\]]", " ", name).strip()
    return (name or "Sheet")[:31]


def write_kv_section(ws, start_row: int, title: str, pairs: Sequence[tuple[str, object]]) -> int:
    ws.cell(start_row, 1, title)
    ws.cell(start_row, 1).font = WHITE_FONT
    ws.cell(start_row, 1).fill = HEADER_FILL
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=4)
    r = start_row + 1
    for key, val in pairs:
        ws.cell(r, 1, key).font = BOLD_FONT
        ws.cell(r, 2, _xlsx_safe(val))
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        ws.cell(r, 2).alignment = Alignment(wrap_text=True, vertical="top")
        r += 1
    return r + 1


def build_workbook(output_path: Path, audit: SinglePoemAudit, spec: RunSpec,
                   metric_summary: pd.DataFrame, sens_summary: pd.DataFrame, sens_variants: pd.DataFrame,
                   line_metrics: pd.DataFrame, stanza_metrics: pd.DataFrame,
                   rolling_summary: pd.DataFrame, rolling_windows: pd.DataFrame,
                   contributors: pd.DataFrame, line_infl: pd.DataFrame, stanza_infl: pd.DataFrame,
                   structure_sound: pd.DataFrame, readability: pd.DataFrame, additional: pd.DataFrame):
    wb = Workbook()
    ws = wb.active
    ws.title = "00 Read Me"
    readme = [
        ("Purpose", "Research-facing post-analysis of one VerseVAD Complete Audit. VerseVAD remains the authoritative measurement layer; this workbook reorganizes exported evidence and adds clearly labeled derived analyses."),
        ("Primary profile", PROFILE_LABELS[spec.primary_profile]),
        ("Primary VAD resource", VAD_RESOURCES.get(spec.primary_vad_resource, spec.primary_vad_resource)),
        ("Sensitivity", "Whole-poem sensitivity uses the authoritative values already exported in 00_START_HERE/profile_comparison.csv. Raw ranges are interpreted only within the same metric/scale."),
        ("Rolling windows", f"Overlapping windows of {spec.rolling_window_size} eligible lexical tokens, step {spec.rolling_step}. A multiword lexical match is retained when any constituent token falls in the current window. These are descriptive trajectories, not detected poetic turns or causal change points." if spec.rolling_windows else "Not requested."),
        ("Line/stanza influence", "Leave-one-unit-out recalculation of whole-poem continuous lexical means under the primary profile. Influence magnitudes are meaningful within a metric; do not compare raw delta sizes across unrelated scales." if spec.influence_analysis else "Not requested."),
        ("Lexical contributors", "Continuous contributor ranking combines distance from the poem mean with effective token/type weight. Binary emotion association contributors are counts/locations only and are never given invented strength scores."),
        ("VAD caution", "Normative lexical VAD ratings describe matched vocabulary. They are not declarations of a poem's emotion, speaker psychology, authorial intent, or reader response."),
        ("PoetryID caution", "Archetype/neighbor evidence is interpretive lexical profiling. Affinity and nearest-centroid evidence are not probabilities."),
        ("Audit schema", audit.schema_version if audit.schema_version is not None else "unknown"),
        ("Toolkit version", __version__),
    ]
    ws["A1"] = "VerseVAD Single-Poem Post-Analysis"
    ws["A1"].font = Font(bold=True, size=16)
    ws.merge_cells("A1:D1")
    r = 3
    for key, val in readme:
        ws.cell(r, 1, key).font = BOLD_FONT
        ws.cell(r, 2, val)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        ws.cell(r, 2).alignment = Alignment(wrap_text=True, vertical="top")
        r += 1
    for c, w in {"A": 24, "B": 35, "C": 22, "D": 22}.items():
        ws.column_dimensions[c].width = w

    # Overview
    ov = wb.create_sheet("01 Poem Overview")
    ov["A1"] = audit.title
    ov["A1"].font = Font(bold=True, size=16)
    ov.merge_cells("A1:D1")
    r = write_kv_section(ov, 3, "Poem & Run", [
        ("Title", audit.title), ("Author", audit.author),
        ("Primary profile", PROFILE_LABELS[spec.primary_profile]),
        ("Primary VAD resource", VAD_RESOURCES.get(spec.primary_vad_resource, spec.primary_vad_resource)),
        ("Selected lexical metrics", len(metric_summary)),
    ])
    # Core headline metrics
    ov.cell(r, 1, "Headline Lexical Measurements")
    ov.cell(r, 1).font = WHITE_FONT
    ov.cell(r, 1).fill = HEADER_FILL
    ov.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    r += 1
    headers = ["Metric", "Value", "Unit", "Coverage"]
    for j, h in enumerate(headers, 1):
        ov.cell(r, j, h).font = BOLD_FONT
        ov.cell(r, j).fill = SUBHEADER_FILL
    r += 1
    for _, x in metric_summary.head(25).iterrows():
        ov.cell(r, 1, x["metric"])
        ov.cell(r, 2, _xlsx_safe(x["value"]))
        ov.cell(r, 3, x["unit"])
        ov.cell(r, 4, _xlsx_safe(x["coverage"]))
        if pd.notna(x["coverage"]):
            ov.cell(r, 4).number_format = "0.0%"
        r += 1
    r += 1
    if not sens_summary.empty:
        strongest = sens_summary.loc[sens_summary["maximum_absolute_change"].idxmax()]
        r = write_kv_section(ov, r, "Largest Methodological Sensitivity Observed", [
            ("Metric", strongest["metric"]),
            ("Range", strongest["maximum_absolute_change"]),
            ("Lowest", f"{strongest['lowest_value']:.6g} · {strongest['lowest_variant']}"),
            ("Highest", f"{strongest['highest_value']:.6g} · {strongest['highest_variant']}"),
        ])
    for label, infl in [("Line", line_infl), ("Stanza", stanza_infl)]:
        summ = influence_summary(infl)
        if not summ.empty:
            strongest = summ.loc[summ["largest_absolute_change"].idxmax()]
            r = write_kv_section(ov, r, f"Largest {label}-Removal Influence Observed", [
                ("Metric", strongest["metric"]),
                (f"{label}", int(strongest["most_influential_ordinal"])),
                ("Absolute change", strongest["largest_absolute_change"]),
                ("Interpretation", "Compare this delta only with other removals for the same metric/scale."),
            ])
    _autosize(ov, 55)

    write_df_sheet(wb, "02 Lexical Profile", metric_summary)
    write_df_sheet(wb, "03 Sensitivity Summary", sens_summary)
    write_df_sheet(wb, "04 Sensitivity Variants", sens_variants)
    write_df_sheet(wb, "05 Line Dynamics", line_metrics)
    write_df_sheet(wb, "06 Stanza Dynamics", stanza_metrics)
    write_df_sheet(wb, "07 Rolling Summary", rolling_summary)
    write_df_sheet(wb, "08 Rolling Windows", rolling_windows)
    write_df_sheet(wb, "09 Contributors", contributors)
    write_df_sheet(wb, "10 Line Influence", line_infl)
    write_df_sheet(wb, "11 Stanza Influence", stanza_infl)
    write_df_sheet(wb, "12 Structure Sound Form", structure_sound)
    write_df_sheet(wb, "13 Readability", readability)
    write_df_sheet(wb, "14 Coverage Evidence", audit.coverage.copy())
    write_df_sheet(wb, "15 Additional Modules", additional)

    meta = pd.DataFrame([
        {"field": "source_zip", "value": str(audit.path)},
        {"field": "title", "value": audit.title},
        {"field": "author", "value": audit.author},
        {"field": "audit_schema_version", "value": audit.schema_version},
        {"field": "single_poem_analysis_version", "value": __version__},
        {"field": "primary_profile", "value": spec.primary_profile},
        {"field": "primary_vad_resource", "value": spec.primary_vad_resource},
        {"field": "generated_at", "value": datetime.now().isoformat(timespec="seconds")},
    ])
    write_df_sheet(wb, "Run Metadata", meta)
    wb.save(output_path)


def run_analysis(audit: SinglePoemAudit, root: Path, args) -> Path:
    print(f"\nVerseVAD Single-Poem Analysis v{__version__}")
    print("=" * 44)
    print(f"Title:  {audit.title}")
    if audit.author:
        print(f"Author: {audit.author}")
    if audit.schema_version is not None:
        print(f"Audit schema: {audit.schema_version}")

    print("\nAvailable evidence")
    for label, ok in audit.evidence_availability().items():
        print(f"  {'✓' if ok else '–'} {label}")

    primary_profile = choose_profile(args)
    primary_vad = choose_vad_resource(audit, args)

    if args.quick:
        selected = audit.core_metrics(primary_vad)
    else:
        use_core = prompt_yes_no("\nUse the recommended core lexical metrics?", True)
        selected = audit.core_metrics(primary_vad) if use_core else []
        if selected:
            print(f"  Added {len(selected)} core metrics: VAD, Concreteness, Frequency, AoA, and Lancaster sensorimotor domains.")
        if prompt_yes_no("Add additional profile-aware metrics?", False):
            selected = metric_search_loop(audit, selected, primary_vad)
    if not selected:
        raise SystemExit("No metrics selected.")

    # Sensitivity choices
    if args.sensitivity_profiles:
        if args.sensitivity_profiles.upper() == "A":
            sensitivity_profiles = PROFILE_ORDER.copy()
        else:
            inds = parse_multi_selection(args.sensitivity_profiles, len(PROFILE_ORDER), allow_all=True)
            sensitivity_profiles = [PROFILE_ORDER[i - 1] for i in inds]
    elif args.quick:
        sensitivity_profiles = PROFILE_ORDER.copy()
    else:
        print("\nSensitivity profiles")
        for i, pid in enumerate(PROFILE_ORDER, 1):
            print(f"[{i}] {PROFILE_LABELS[pid]}")
        inds = prompt_multi("Select profiles [A]: ", len(PROFILE_ORDER), default=list(range(1, 7)), allow_all=True)
        sensitivity_profiles = [PROFILE_ORDER[i - 1] for i in inds]

    vad_resources_available = audit.vad_resources()
    if primary_vad and any(m.module_id == "vad" for m in selected):
        ids = [x[0] for x in vad_resources_available]
        if args.sensitivity_vad:
            if args.sensitivity_vad.upper() == "A":
                sensitivity_vad = ids
            else:
                inds = parse_multi_selection(args.sensitivity_vad, len(ids), allow_all=True)
                sensitivity_vad = [ids[i - 1] for i in inds]
        elif args.quick:
            sensitivity_vad = ids
        else:
            print("\nVAD resources for sensitivity analysis")
            for i, (rid, label) in enumerate(vad_resources_available, 1):
                print(f"[{i}] {label} ({rid})")
            inds = prompt_multi("Select resources [A]: ", len(ids), default=list(range(1, len(ids) + 1)), allow_all=True)
            sensitivity_vad = [ids[i - 1] for i in inds]
    else:
        sensitivity_vad = []

    if args.quick:
        rolling = not args.no_rolling
        window_size = args.window_size or 25
        step = args.window_step or 5
        influence = not args.no_influence
        top_contrib = args.top_contributors or 20
    else:
        rolling = False if args.no_rolling else prompt_yes_no("\nRun rolling eligible-token window analysis?", True)
        window_size = args.window_size or (prompt_int("Rolling window size in eligible lexical tokens", 25, minimum=5) if rolling else 25)
        step = args.window_step or (prompt_int("Rolling window step", 5, minimum=1) if rolling else 5)
        influence = False if args.no_influence else prompt_yes_no("Run line/stanza leave-one-unit-out influence analysis?", True)
        top_contrib = args.top_contributors or prompt_int("Top lexical contributors retained per metric", 20, minimum=5, maximum=100)

    spec = RunSpec(
        source_zip=str(audit.path), primary_profile=primary_profile,
        primary_vad_resource=primary_vad,
        selected_metrics=[asdict(x) for x in selected],
        sensitivity_profiles=sensitivity_profiles,
        sensitivity_vad_resources=sensitivity_vad,
        rolling_windows=rolling, rolling_window_size=window_size, rolling_step=step,
        influence_analysis=influence, top_contributors_per_metric=top_contrib,
    )

    print("\nAnalysis configuration")
    print(f"  Primary profile: {PROFILE_LABELS[primary_profile]}")
    if primary_vad:
        print(f"  Primary VAD:     {VAD_RESOURCES.get(primary_vad, primary_vad)}")
    print(f"  Metrics:         {len(selected)}")
    print(f"  Sensitivity:     {len(sensitivity_profiles)} profile(s)" + (f" × {len(sensitivity_vad)} VAD resource(s)" if sensitivity_vad else ""))
    print(f"  Rolling windows: {'yes' if rolling else 'no'}" + (f" ({window_size} tokens, step {step})" if rolling else ""))
    print(f"  Influence:       {'yes' if influence else 'no'}")
    if not args.quick and not prompt_yes_no("\nRun analysis?", True):
        raise SystemExit("Cancelled.")

    metric_summary = build_metric_summary(audit, selected, primary_profile)
    sens_summary, sens_variants = build_sensitivity(audit, selected, primary_profile, sensitivity_profiles, sensitivity_vad)
    line_metrics = build_unit_dynamics(audit, selected, primary_profile, "line")
    stanza_metrics = build_unit_dynamics(audit, selected, primary_profile, "stanza")
    rolling_windows, rolling_summary = (build_rolling_windows(audit, selected, primary_profile, window_size, step) if rolling else (pd.DataFrame(), pd.DataFrame()))
    contributors = build_contributors(audit, selected, primary_profile, top_contrib)
    line_infl = build_influence(audit, selected, primary_profile, "line") if influence else pd.DataFrame()
    stanza_infl = build_influence(audit, selected, primary_profile, "stanza") if influence else pd.DataFrame()
    structure_sound = build_structure_sound_form(audit)
    readability = build_readability(audit)
    additional = build_additional_module_summary(audit, primary_profile, primary_vad)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_root = Path(args.output).expanduser().resolve() if args.output else root / "exports" / "single_poem"
    run_dir = out_root / f"{slugify(audit.title)}_{timestamp}"
    run_dir.mkdir(parents=True, exist_ok=False)

    outputs = {
        "metric_summary.csv": metric_summary,
        "sensitivity_summary.csv": sens_summary,
        "sensitivity_variants.csv": sens_variants,
        "line_metrics.csv": line_metrics,
        "stanza_metrics.csv": stanza_metrics,
        "rolling_summary.csv": rolling_summary,
        "rolling_windows.csv": rolling_windows,
        "lexical_contributors.csv": contributors,
        "line_influence.csv": line_infl,
        "stanza_influence.csv": stanza_infl,
        "coverage_summary.csv": audit.coverage.copy(),
        "structure_sound_form.csv": structure_sound,
        "readability_summary.csv": readability,
        "additional_module_summary.csv": additional,
    }
    for name, df in outputs.items():
        save_csv(df, run_dir / name)

    with open(run_dir / "analysis_spec.json", "w", encoding="utf-8") as fh:
        json.dump(asdict(spec), fh, indent=2, ensure_ascii=False)

    adapter_checks = []
    for m in selected:
        if m.module_id in DERIVABLE_CONTINUOUS_MODULES:
            ok, derived, exported = audit.validate_observation_adapter(m, primary_profile)
            adapter_checks.append({
                "metric": m.metric, "resource": m.source, "validated": ok,
                "derived_mean": None if not np.isfinite(derived) else derived,
                "exported_mean": None if not np.isfinite(exported) else exported,
                "absolute_error": None if not (np.isfinite(derived) and np.isfinite(exported)) else abs(derived - exported),
            })
    metadata = {
        "single_poem_analysis_version": __version__,
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "source_zip": str(audit.path),
        "source_zip_bytes": audit.path.stat().st_size,
        "audit_schema_version": audit.schema_version,
        "title": audit.title,
        "author": audit.author,
        "adapter_validation": adapter_checks,
        "notes": [
            "Whole-poem profile/sensitivity values are taken from VerseVAD's exported profile_comparison.csv.",
            "Derived line/stanza/rolling/influence calculations are emitted only when the observation adapter reproduces the authoritative primary-profile mean within numerical tolerance.",
            "Raw influence and sensitivity ranges are comparable within a metric, not across unrelated scales.",
        ],
    }
    with open(run_dir / "analysis_metadata.json", "w", encoding="utf-8") as fh:
        json.dump(metadata, fh, indent=2, ensure_ascii=False)

    workbook = run_dir / "single_poem_analysis.xlsx"
    build_workbook(
        workbook, audit, spec, metric_summary, sens_summary, sens_variants,
        line_metrics, stanza_metrics, rolling_summary, rolling_windows,
        contributors, line_infl, stanza_infl, structure_sound, readability, additional,
    )

    print("\nComplete.")
    print(f"Output folder:\n  {run_dir}")
    print("\nOpen first:\n  single_poem_analysis.xlsx")
    print(f"\nPrimary-profile metrics: {len(metric_summary)}")
    print(f"Sensitivity variants:     {len(sens_variants)}")
    print(f"Line metric rows:         {len(line_metrics)}")
    print(f"Stanza metric rows:       {len(stanza_metrics)}")
    print(f"Rolling-window rows:      {len(rolling_windows)}")
    print(f"Contributor rows:         {len(contributors)}")
    print(f"Line influence rows:      {len(line_infl)}")
    print(f"Stanza influence rows:    {len(stanza_infl)}")
    failed_checks = [x for x in adapter_checks if not x["validated"]]
    if failed_checks:
        print(f"\nNote: {len(failed_checks)} selected metric adapter(s) did not exactly reproduce the exported primary-profile mean; derived dynamics/influence for those metrics were skipped. See analysis_metadata.json.")
    return run_dir


def build_arg_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(description="Post-analyze a VerseVAD single-poem Complete Audit ZIP.")
    p.add_argument("--source", help="Path to a single-poem VerseVAD Complete Audit ZIP.")
    p.add_argument("--output", help="Output root. Default: exports/single_poem/")
    p.add_argument("--version", action="version", version=f"single.py {__version__}")
    p.add_argument("--quick", action="store_true", help="Run non-interactively with recommended defaults; useful for testing/reproducible batch runs.")
    p.add_argument("--primary-profile", help="Primary profile ID or 1-6.")
    p.add_argument("--primary-vad", help="Primary VAD resource ID.")
    p.add_argument("--sensitivity-profiles", help="A or selection syntax such as 1,3,5-6.")
    p.add_argument("--sensitivity-vad", help="A or selection syntax over available VAD resources.")
    p.add_argument("--window-size", type=int, help="Rolling window size in eligible lexical tokens.")
    p.add_argument("--window-step", type=int, help="Rolling window step.")
    p.add_argument("--no-rolling", action="store_true")
    p.add_argument("--no-influence", action="store_true")
    p.add_argument("--top-contributors", type=int)
    return p


def main(argv: Optional[Sequence[str]] = None) -> int:
    configure_console_encoding()
    args = build_arg_parser().parse_args(argv)
    script_path = Path(__file__).resolve()
    try:
        root, source = choose_source(args, script_path)
        audit = SinglePoemAudit(source)
        if audit.schema_version not in {None, 2, 3}:
            print(
                "Warning: this script supports VerseVAD export schemas 2 and 3; "
                f"the archive reports v{audit.schema_version}."
            )
        run_analysis(audit, root, args)
        return 0
    except (AuditError, KeyboardInterrupt) as exc:
        if isinstance(exc, KeyboardInterrupt):
            print("\nCancelled.")
        else:
            print(f"Error: {exc}", file=sys.stderr)
        return 2


if __name__ == "__main__":
    raise SystemExit(main())
